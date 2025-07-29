import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Tuple
import logging

logger = logging.getLogger(__name__)

class BlankAnalyzer:
    def __init__(self, blank_file: str):
        """Ініціалізація аналізатора бланка"""
        self.blank_file = blank_file
        self.structure = self.analyze_blank_structure()
    
    def analyze_blank_structure(self) -> Dict:
        """Аналіз структури бланка"""
        try:
            workbook = load_workbook(self.blank_file)
            structure = {
                'sheets': [],
                'headers': {},
                'product_columns': {},
                'total_columns': {},
                'bakery_info': {}
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = {
                    'name': sheet_name,
                    'headers': [],
                    'product_rows': [],
                    'total_rows': [],
                    'bakery_info_rows': []
                }
                
                # Аналізуємо кожен рядок
                for row in range(1, sheet.max_row + 1):
                    row_data = []
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value:
                            row_data.append(str(cell_value).strip())
                    
                    if row_data:
                        # Шукаємо заголовки
                        if any(keyword in ' '.join(row_data).upper() for keyword in 
                               ['НАЗВА', 'ПРОДУКТ', 'КІЛЬКІСТЬ', 'ЦІНА', 'СУМА', 'РАЗОМ']):
                            sheet_info['headers'].append({
                                'row': row,
                                'data': row_data
                            })
                        
                        # Шукаємо інформацію про пекарню
                        if any(keyword in ' '.join(row_data).upper() for keyword in 
                               ['ПЕКАРНЯ', 'ТОВ', 'ПП', 'ФОП', 'ПРОДАВЕЦЬ']):
                            sheet_info['bakery_info_rows'].append({
                                'row': row,
                                'data': row_data
                            })
                        
                        # Шукаємо рядки з продуктами (містять цифри)
                        if any(char.isdigit() for item in row_data for char in str(item)):
                            sheet_info['product_rows'].append({
                                'row': row,
                                'data': row_data
                            })
                        
                        # Шукаємо підсумкові рядки
                        if any(keyword in ' '.join(row_data).upper() for keyword in 
                               ['РАЗОМ', 'ВСЬОГО', 'СУМА', 'ПІДСУМОК']):
                            sheet_info['total_rows'].append({
                                'row': row,
                                'data': row_data
                            })
                
                structure['sheets'].append(sheet_info)
            
            logger.info(f"Проаналізовано бланк: {len(structure['sheets'])} аркушів")
            return structure
            
        except Exception as e:
            logger.error(f"Помилка аналізу бланка: {e}")
            return {}
    
    def get_expected_columns(self) -> List[str]:
        """Отримання очікуваних колонок з бланка"""
        expected_columns = []
        
        for sheet in self.structure.get('sheets', []):
            for header in sheet.get('headers', []):
                for item in header['data']:
                    if any(keyword in item.upper() for keyword in 
                           ['НАЗВА', 'ПРОДУКТ', 'КІЛЬКІСТЬ', 'ЦІНА', 'СУМА', 'РАЗОМ']):
                        expected_columns.append(item)
        
        return list(set(expected_columns))
    
    def get_bakery_name_patterns(self) -> List[str]:
        """Отримання патернів для назв пекарень з бланка"""
        patterns = []
        
        for sheet in self.structure.get('sheets', []):
            for row_info in sheet.get('bakery_info_rows', []):
                for item in row_info['data']:
                    if any(keyword in item.upper() for keyword in 
                           ['ПЕКАРНЯ', 'ТОВ', 'ПП', 'ФОП']):
                        patterns.append(item)
        
        return list(set(patterns))
    
    def get_product_format_examples(self) -> List[Dict]:
        """Отримання прикладів формату продуктів з бланка"""
        examples = []
        
        for sheet in self.structure.get('sheets', []):
            for row_info in sheet.get('product_rows', []):
                if len(row_info['data']) >= 3:  # Мінімум назва, кількість, ціна
                    example = {
                        'name': row_info['data'][0] if len(row_info['data']) > 0 else '',
                        'quantity': row_info['data'][1] if len(row_info['data']) > 1 else '',
                        'price': row_info['data'][2] if len(row_info['data']) > 2 else '',
                        'total': row_info['data'][3] if len(row_info['data']) > 3 else ''
                    }
                    examples.append(example)
        
        return examples
    
    def generate_ocr_patterns(self) -> Dict:
        """Генерація патернів для OCR на основі бланка"""
        patterns = {
            'bakery_name_patterns': self.get_bakery_name_patterns(),
            'product_patterns': [],
            'column_headers': self.get_expected_columns(),
            'examples': self.get_product_format_examples()
        }
        
        # Генеруємо патерни для продуктів на основі прикладів
        for example in patterns['examples']:
            if example['name'] and example['quantity']:
                # Патерн: назва кількість ціна
                if example['price']:
                    patterns['product_patterns'].append(
                        f"([А-ЯІЇЄ\\w\\s]+)\\s+(\\d+(?:[.,]\\d+)?)\\s+(\\d+(?:[.,]\\d+)?)"
                    )
                # Патерн: назва кількість
                else:
                    patterns['product_patterns'].append(
                        f"([А-ЯІЇЄ\\w\\s]+)\\s+(\\d+(?:[.,]\\d+)?)"
                    )
        
        return patterns
    
    def print_analysis_report(self):
        """Виведення звіту про аналіз бланка"""
        print("📋 АНАЛІЗ БЛАНКА НАКЛАДНИХ")
        print("=" * 50)
        
        print(f"📄 Файл: {self.blank_file}")
        print(f"📊 Аркушів: {len(self.structure.get('sheets', []))}")
        
        for i, sheet in enumerate(self.structure.get('sheets', []), 1):
            print(f"\n📋 Аркуш {i}: {sheet['name']}")
            print(f"   Заголовків: {len(sheet.get('headers', []))}")
            print(f"   Рядків з продуктами: {len(sheet.get('product_rows', []))}")
            print(f"   Рядків з інфо про пекарню: {len(sheet.get('bakery_info_rows', []))}")
            print(f"   Підсумкових рядків: {len(sheet.get('total_rows', []))}")
        
        patterns = self.generate_ocr_patterns()
        print(f"\n🎯 ЗНАЙДЕНІ ПАТЕРНИ:")
        print(f"   Назви пекарень: {len(patterns['bakery_name_patterns'])}")
        print(f"   Патерни продуктів: {len(patterns['product_patterns'])}")
        print(f"   Заголовки колонок: {len(patterns['column_headers'])}")
        
        if patterns['examples']:
            print(f"\n📝 ПРИКЛАДИ ФОРМАТУ:")
            for i, example in enumerate(patterns['examples'][:3], 1):
                print(f"   {i}. {example['name']} | {example['quantity']} | {example['price']} | {example['total']}")

def main():
    """Тестування аналізатора бланка"""
    blank_file = "бланк для випічки з новинками.xlsx"
    
    if not os.path.exists(blank_file):
        print(f"❌ Файл {blank_file} не знайдено")
        return
    
    analyzer = BlankAnalyzer(blank_file)
    analyzer.print_analysis_report()
    
    # Зберігаємо патерни для використання в OCR
    patterns = analyzer.generate_ocr_patterns()
    
    with open('blank_patterns.json', 'w', encoding='utf-8') as f:
        json.dump(patterns, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Патерни збережено в blank_patterns.json")

if __name__ == "__main__":
    import os
    import json
    main() 