import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict
import os
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        """Ініціалізація генератора Excel"""
        self.output_dir = "excel_reports"
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"Створено папку: {self.output_dir}")
    
    def create_excel(self, bakery_name: str, products: List[Dict], date: str = None) -> str:
        """Створення Excel файлу для накладної"""
        if not date:
            date = datetime.now().strftime("%d.%m")
        
        # Очищаємо назву пекарні для файлу
        safe_bakery_name = self.sanitize_filename(bakery_name) if bakery_name else "Невідома_пекарня"
        filename = f"{safe_bakery_name}_{date}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Створюємо DataFrame
        df = self.create_dataframe(products)
        
        # Створюємо Excel файл
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Накладна"
        
        # Додаємо заголовок
        self.add_header(worksheet, bakery_name, date)
        
        # Додаємо дані
        self.add_data(worksheet, df)
        
        # Форматуємо
        self.format_worksheet(worksheet, df)
        
        # Зберігаємо
        workbook.save(filepath)
        logger.info(f"Створено Excel файл: {filepath}")
        
        return filepath
    
    def create_dataframe(self, products: List[Dict]) -> pd.DataFrame:
        """Створення DataFrame з продуктами"""
        if not products:
            return pd.DataFrame(columns=['№', 'Назва продукту', 'Кількість', 'Ціна', 'Сума'])
        
        data = []
        for i, product in enumerate(products, 1):
            data.append({
                '№': i,
                'Назва продукту': product.get('name', ''),
                'Кількість': product.get('quantity', 0),
                'Ціна': product.get('price', 0),
                'Сума': product.get('total', 0)
            })
        
        return pd.DataFrame(data)
    
    def add_header(self, worksheet, bakery_name: str, date: str):
        """Додавання заголовка"""
        # Назва пекарні
        worksheet['A1'] = f"Пекарня: {bakery_name if bakery_name else 'Невідома'}"
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:E1')
        
        # Дата
        worksheet['A2'] = f"Дата: {date}"
        worksheet['A2'].font = Font(bold=True)
        worksheet.merge_cells('A2:E2')
        
        # Порожній рядок
        worksheet['A3'] = ""
        
        # Заголовки таблиці
        headers = ['№', 'Назва продукту', 'Кількість', 'Ціна', 'Сума']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def add_data(self, worksheet, df: pd.DataFrame):
        """Додавання даних"""
        if df.empty:
            return
        
        # Додаємо дані з DataFrame
        for r in dataframe_to_rows(df, index=False, header=False):
            worksheet.append(r)
        
        # Додаємо підсумковий рядок
        total_row = len(df) + 5
        worksheet[f'A{total_row}'] = "РАЗОМ:"
        worksheet[f'A{total_row}'].font = Font(bold=True)
        worksheet.merge_cells(f'A{total_row}:C{total_row}')
        
        # Підраховуємо підсумки
        total_quantity = df['Кількість'].sum()
        total_amount = df['Сума'].sum()
        
        worksheet[f'D{total_row}'] = total_quantity
        worksheet[f'E{total_row}'] = total_amount
        worksheet[f'D{total_row}'].font = Font(bold=True)
        worksheet[f'E{total_row}'].font = Font(bold=True)
    
    def format_worksheet(self, worksheet, df: pd.DataFrame):
        """Форматування робочого аркуша"""
        # Встановлюємо ширину стовпців
        column_widths = [5, 40, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Форматуємо числа
        for row in range(5, len(df) + 6):
            # Кількість
            cell = worksheet[f'C{row}']
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            
            # Ціна
            cell = worksheet[f'D{row}']
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            
            # Сума
            cell = worksheet[f'E{row}']
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        
        # Додаємо рамки
        self.add_borders(worksheet, len(df) + 4)
        
        # Центруємо заголовки
        for col in range(1, 6):
            cell = worksheet.cell(row=4, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def add_borders(self, worksheet, last_row: int):
        """Додавання рамок"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(4, last_row + 1):
            for col in range(1, 6):
                worksheet.cell(row=row, column=col).border = thin_border
    
    def sanitize_filename(self, filename: str) -> str:
        """Очищення назви файлу від недопустимих символів"""
        # Замінюємо недопустимі символи
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        
        # Обмежуємо довжину
        if len(filename) > 50:
            filename = filename[:50]
        
        return filename.strip()
    
    def create_multiple_sheets_excel(self, invoices_data: List[Dict]) -> str:
        """Створення Excel файлу з кількома аркушами для різних накладних"""
        if not invoices_data:
            return None
        
        date = datetime.now().strftime("%d.%m")
        filename = f"Всі_накладні_{date}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        workbook = openpyxl.Workbook()
        
        # Видаляємо стандартний аркуш
        workbook.remove(workbook.active)
        
        for i, invoice in enumerate(invoices_data):
            bakery_name = invoice.get('bakery_name', f'Накладна_{i+1}')
            products = invoice.get('products', [])
            
            # Створюємо аркуш
            worksheet = workbook.create_sheet(title=f"Накладна_{i+1}")
            
            # Додаємо дані
            df = self.create_dataframe(products)
            self.add_header(worksheet, bakery_name, date)
            self.add_data(worksheet, df)
            self.format_worksheet(worksheet, df)
        
        workbook.save(filepath)
        logger.info(f"Створено Excel файл з кількома аркушами: {filepath}")
        
        return filepath 